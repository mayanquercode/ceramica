import openpyxl
import json
import os

# Función para leer un archivo JSON
def read_json_file(file_path):
    """                                               Reads a JSON file, converts it for Python usage, and returns its content.                       
    Args:
        file_path (str): The path to the JSON file.
    Returns:                                              dict: The transformed JSON content as a Python dictionary.
    """
    try:
        with open(file_path, 'r') as file:
            json_data = json.load(file)
        return json_data
    except FileNotFoundError:
        print("Error: The file was not found. Please check the path and try again.")
    except json.JSONDecodeError:
        print("Error: The file is not a valid JSON. Please provide a proper JSON file.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

# Función para verificar y obtener una hoja específica de un archivo Excel
def read_excel_sheet(excel_path, sheet_name):
    """
    Verifica si el archivo Excel existe y si contiene una hoja específica.

    Args:
        excel_path (str): Ruta al archivo Excel.
        sheet_name (str): Nombre de la hoja a buscar.

    Returns:
        openpyxl.worksheet.worksheet.Worksheet: La hoja de trabajo si existe.

    Raises:
        FileNotFoundError: Si el archivo Excel no existe.
        ValueError: Si la hoja no existe en el archivo Excel.
    """
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo Excel.")
        return workbook[sheet_name]
    except FileNotFoundError:
        raise FileNotFoundError(f"El archivo '{excel_path}' no fue encontrado.")
    except Exception as e:
        raise Exception(f"Error al intentar leer el archivo Excel: {e}")

# Función para extraer los datos con la estructura especificada
def extract_data_from_sheet(sheet, categories):
    """
    Extrae datos de una hoja de Excel, filtra según las categorías y genera una lista de objetos (diccionarios)
    con los campos: code, name, stock y category.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): La hoja de trabajo de Excel.
        categories (list): Lista de categorías válidas para filtrar.

    Returns:
        list: Una lista de diccionarios con los datos extraídos.
    """
    data = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # Ignorar encabezados
        code = row[5].value  # Columna F
        name = row[6].value  # Columna G
        stock = row[8].value  # Columna I
        category = row[2].value  # Columna C

        if code and name and category and category in categories:  # Verificar que la categoría sea válida
            item = {
                "code": code,
                "name": name,
                "stock": stock if stock is not None else 0,  # Stock predeterminado a 0 si es nulo
                "category": category
            }
            data.append(item)
    return data

# Función para guardar los datos en un archivo JSON
def save_to_json_file(data, output_path):
    """
    Guarda los datos en un archivo JSON.

    Args:
        data (list): Los datos a guardar.
        output_path (str): Ruta donde se creará el archivo JSON.
    """
    try:
        with open(output_path, "w", encoding="utf-8") as json_file:
            json.dump(data, json_file, indent=4, ensure_ascii=False)
        print(f"Archivo JSON guardado en: {output_path}")
    except Exception as e:
        print(f"Error al guardar el archivo JSON: {e}")

# Ejecución principal
if __name__ == "__main__":
    excel_path = "/data/data/com.termux/files/home/storage/downloads/RdlReporteStock.xlsx"
    inventory_path = "/data/data/com.termux/files/home/projects/ceramica/backend/data/store.json"
    products_json_path = "/data/data/com.termux/files/home/projects/ceramica/backend/data/inventory.json"
    sheet_name = "RdlReporteStock"

    try:
        # Leer el archivo JSON
        inventory = read_json_file(inventory_path)
        categories = [cat["name"] for cat in inventory.get('categories', {}).values()]

        # Leer y verificar la hoja del archivo Excel
        sheet = read_excel_sheet(excel_path, sheet_name)
        print(f"La hoja '{sheet_name}' fue encontrada en el archivo Excel.")

        # Extraer y filtrar los datos de la hoja
        extracted_data = extract_data_from_sheet(sheet, categories)
        print(f"Se extrajeron {len(extracted_data)} filas con categorías válidas.")

        # Guardar los datos filtrados en un archivo JSON
        save_to_json_file(extracted_data, products_json_path)

    except Exception as e:
        print(f"Error: {e}")
